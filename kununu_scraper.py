#!/usr/bin/env python3
"""
kununu.com Web Scraper
Scrapet Unternehmensdaten von kununu.com und exportiert sie als CSV/XLSX.

Nutzung:
    python kununu_scraper.py --hilfe

Hinweis: Bitte beachte die Nutzungsbedingungen von kununu.com.
         Scrape verantwortungsvoll und mit angemessenen Pausen.
"""

import argparse
import csv
import logging
import random
import re
import sys
import time
from dataclasses import dataclass, fields, asdict
from pathlib import Path
from urllib.parse import urlencode, urljoin

import requests
from bs4 import BeautifulSoup

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_VERFUEGBAR = True
except ImportError:
    PLAYWRIGHT_VERFUEGBAR = False

# Logging konfigurieren
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

BASE_URL = "https://www.kununu.com"
SEARCH_URL = f"{BASE_URL}/de/search"

# Realistische Browser-Header
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}


@dataclass
class Unternehmen:
    """Datenklasse für ein Unternehmen auf kununu."""

    name: str = ""
    score: str = ""
    anzahl_bewertungen: str = ""
    standort: str = ""
    branche: str = ""
    profil_url: str = ""
    weiterempfehlung: str = ""
    # Detail-Felder (bei --details)
    gehaltszufriedenheit: str = ""
    karriere_weiterbildung: str = ""
    unternehmenskultur: str = ""
    arbeitsumgebung: str = ""
    vielfalt: str = ""
    kollegenzusammenhalt: str = ""
    kommunikation: str = ""
    top_company: str = ""


def pause(min_sec: float = 1.5, max_sec: float = 4.0) -> None:
    """Zufällige Pause um den Server nicht zu überlasten."""
    wartezeit = random.uniform(min_sec, max_sec)
    time.sleep(wartezeit)


def seite_abrufen(url: str, session: requests.Session) -> BeautifulSoup | None:
    """Ruft eine Seite ab und gibt ein BeautifulSoup-Objekt zurück (requests)."""
    try:
        response = session.get(url, headers=HEADERS, timeout=30)
        response.raise_for_status()
        return BeautifulSoup(response.text, "html.parser")
    except requests.RequestException as e:
        log.error("Fehler beim Abrufen von %s: %s", url, e)
        return None


def seite_abrufen_browser(url: str, page) -> BeautifulSoup | None:
    """Ruft eine Seite mit Playwright ab (rendert JavaScript)."""
    try:
        page.goto(url, wait_until="networkidle", timeout=30000)
        # Cookie-Banner schließen falls vorhanden
        try:
            accept_btn = page.locator("button:has-text('Akzeptieren')").first
            if accept_btn.is_visible(timeout=2000):
                accept_btn.click()
                page.wait_for_timeout(1000)
        except Exception:
            pass
        # Warten bis Inhalte geladen sind
        page.wait_for_timeout(2000)
        html = page.content()
        return BeautifulSoup(html, "html.parser")
    except Exception as e:
        log.error("Browser-Fehler beim Abrufen von %s: %s", url, e)
        return None


def text_bereinigen(text: str) -> str:
    """Bereinigt Text von überflüssigen Leerzeichen."""
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def score_extrahieren(text: str) -> str:
    """Extrahiert einen numerischen Score aus Text."""
    match = re.search(r"(\d[.,]\d)", text)
    if match:
        return match.group(1)
    return ""


def unternehmen_von_suchseite(soup: BeautifulSoup) -> list[Unternehmen]:
    """Extrahiert Unternehmensdaten aus einer Suchergebnis-Seite."""
    ergebnisse = []

    # Suche nach Unternehmens-Links im Format /de/firmenname
    profil_links = soup.find_all("a", href=re.compile(r"^/de/[\w-]+$"))

    # Verarbeite die Seite blockweise - jedes Unternehmen hat einen Link
    # zum Profil mit "Go to ... profile" im Text
    go_to_links = soup.find_all("a", attrs={"href": re.compile(r"^/de/[\w-]+/?$")})

    gesehen = set()
    for link in go_to_links:
        href = link.get("href", "")
        # Normalerweise ist href /de/firmenname
        if not href or href in gesehen:
            continue
        # Filter nur Unternehmensprofile (nicht /de/search, /de/login, etc.)
        skip_paths = {
            "/de/search", "/de/login", "/de/gehalt", "/de/jobs",
            "/de/beste-arbeitgeber", "/de/user", "/de/insights",
        }
        if href.rstrip("/") in skip_paths or "/kommentare" in href:
            continue

        gesehen.add(href)
        u = Unternehmen()
        u.profil_url = urljoin(BASE_URL, href)

        # Versuche den Kontext um den Link herum zu finden
        # Der Elterncontainer enthält Name, Score, Ort, Branche
        container = link.find_parent(["article", "div", "section"])
        if not container:
            container = link.parent
            if container:
                container = container.parent

        if container:
            text_content = text_bereinigen(container.get_text())

            # Name extrahieren: bevorzuge aria-label des Links (enthält "Go to X profile")
            aria = link.get("aria-label", "") or ""
            if "Go to" in aria and "profile" in aria:
                u.name = aria.replace("Go to ", "").replace(" profile", "").strip()
            else:
                # Fallback: h-Element oder span im Container
                name_elem = container.find(
                    ["h2", "h3", "h4", "span"],
                    string=re.compile(r".{3,}"),
                )
                if name_elem:
                    kandidat = text_bereinigen(name_elem.get_text())
                    # "Top" als alleinigen Namen ignorieren
                    if kandidat and kandidat.lower() not in ("top", "top company"):
                        u.name = kandidat
                if not u.name:
                    u.name = href.split("/")[-1].replace("-", " ").title()

            # Score extrahieren
            score_text = re.search(r"(\d[.,]\d)\s*(?:★|⭐|rating|star|Sterne)?", text_content)
            if score_text:
                u.score = score_text.group(1)

            # Bewertungsanzahl extrahieren (z.B. "22.931 reviews" oder "8.125 Bewertungen")
            bew_match = re.search(r"(?:has\s+)?([\d][\d.,]*\d)\s*(?:review|Bewertung)", text_content)
            if bew_match:
                u.anzahl_bewertungen = bew_match.group(1)

            # Ort extrahieren (Format: "Stadt, Deutschland")
            ort_match = re.search(r"([\w][\w\s-]+,\s*Deutschland)", text_content)
            if ort_match:
                u.standort = text_bereinigen(ort_match.group(1))

            # Branche - steht oft nach dem Ort
            bekannte_branchen = [
                "Automobil", "Industrie", "IT", "Handel", "Versicherung",
                "Transport/Verkehr/Logistik", "Telekommunikation", "Energie",
                "Personalwesen & -beschaffung", "Dienstleistung", "Banken",
                "Beratung/Consulting", "Gesundheit/Soziales/Pflege",
                "Immobilien", "Bildung", "Internet", "Medien",
                "Maschinenbau", "Elektro/Elektronik", "Chemie",
                "Finanz", "Marketing/Werbung/PR", "Öffentliche Verwaltung",
                "Bauwesen", "Tourismus", "Gastronomie",
            ]
            for branche in bekannte_branchen:
                if branche.lower() in text_content.lower():
                    u.branche = branche
                    break

            # Weiterempfehlung extrahieren
            we_match = re.search(r"(\d+)\s*%\s*Weiterempfehlung", text_content)
            if we_match:
                u.weiterempfehlung = we_match.group(1) + "%"

            # Top Company erkennen
            if "Top" in text_content and "Company" in text_content:
                u.top_company = "Ja"

        # Nur hinzufügen wenn wir mindestens einen Namen haben
        skip_names = {"", "Go to", "Sitemap", "Newsletter", "Tracking",
                      "Impressum", "Datenschutz", "AGB"}
        if u.name and u.name not in skip_names and len(u.name) > 2:
            ergebnisse.append(u)

    return ergebnisse


def details_scrapen(u: Unternehmen, session: requests.Session) -> None:
    """Scrapt die Detailseite eines Unternehmens für zusätzliche Infos."""
    if not u.profil_url:
        return

    log.info("  → Details laden: %s", u.name)
    soup = seite_abrufen(u.profil_url, session)
    if not soup:
        return

    text = soup.get_text()

    # kununu Score und Bewertungsanzahl von der Detailseite
    score_match = re.search(r"(\d[.,]\d)\s*kununu\s*Score", text)
    if score_match:
        u.score = score_match.group(1)

    bew_match = re.search(r"([\d.]+)\s*Bewertung", text)
    if bew_match:
        u.anzahl_bewertungen = bew_match.group(1)

    # Weiterempfehlung
    we_match = re.search(r"(\d+)\s*%\s*Weiterempfehlung", text)
    if we_match:
        u.weiterempfehlung = we_match.group(1) + "%"

    # Gehaltszufriedenheit
    gehalt_match = re.search(r"(\d+)\s*%.*(?:Gehalt|Gehälter).*zufrieden", text)
    if gehalt_match:
        u.gehaltszufriedenheit = gehalt_match.group(1) + "%"

    # Kategorie-Scores
    kategorien = {
        "Karriere & Gehalt": "karriere_weiterbildung",
        "Karriere/Weiterbildung": "karriere_weiterbildung",
        "Unternehmenskultur": "unternehmenskultur",
        "Arbeitsumgebung": "arbeitsumgebung",
        "Vielfalt": "vielfalt",
        "Kollegenzusammenhalt": "kollegenzusammenhalt",
        "Kommunikation": "kommunikation",
    }

    for label, feld in kategorien.items():
        pattern = re.compile(
            rf"({label})\s*(\d[.,]\d)",
            re.IGNORECASE,
        )
        match = pattern.search(text)
        if match:
            setattr(u, feld, match.group(2))
        else:
            # Umgekehrte Reihenfolge: Score vor Label
            pattern2 = re.compile(
                rf"(\d[.,]\d)\s*.*?{re.escape(label)}",
                re.IGNORECASE,
            )
            match2 = pattern2.search(text)
            if match2:
                setattr(u, feld, match2.group(1))

    # Top Company
    if "Top Company" in text:
        u.top_company = "Ja"

    pause(2.0, 5.0)


def suchseite_url(seite: int, branche: str = "", ort: str = "",
                   score_min: str = "") -> str:
    """Baut die URL für eine Suchergebnisseite zusammen."""
    params = {"spo": "0", "page": str(seite)}
    if branche:
        params["industry"] = branche
    if ort:
        params["location"] = ort
    if score_min:
        params["score"] = score_min
    return f"{SEARCH_URL}?{urlencode(params)}"


def scrape(
    max_seiten: int = 3,
    branche: str = "",
    ort: str = "",
    score_min: str = "",
    details: bool = False,
    browser_modus: bool = False,
) -> list[Unternehmen]:
    """Hauptfunktion: Scrapt Unternehmen von kununu.com."""
    alle_unternehmen: list[Unternehmen] = []
    gesehene_urls: set[str] = set()

    if browser_modus:
        if not PLAYWRIGHT_VERFUEGBAR:
            log.error(
                "Playwright ist nicht installiert. Installiere es mit:\n"
                "  pip install playwright && python -m playwright install chromium"
            )
            return []
        return _scrape_mit_browser(
            max_seiten, branche, ort, score_min, details,
        )

    session = requests.Session()
    session.headers.update(HEADERS)

    for seite in range(1, max_seiten + 1):
        url = suchseite_url(seite, branche, ort, score_min)
        log.info("Seite %d/%d: %s", seite, max_seiten, url)

        soup = seite_abrufen(url, session)
        if not soup:
            log.warning("Seite %d konnte nicht geladen werden, überspringe...", seite)
            continue

        unternehmen = unternehmen_von_suchseite(soup)
        log.info("  %d Unternehmen auf Seite %d gefunden", len(unternehmen), seite)

        for u in unternehmen:
            if u.profil_url not in gesehene_urls:
                gesehene_urls.add(u.profil_url)
                alle_unternehmen.append(u)

        if not unternehmen:
            log.info("Keine weiteren Unternehmen gefunden, Abbruch.")
            break

        pause()

    # Optional: Detailseiten scrapen
    if details and alle_unternehmen:
        log.info("Lade Details für %d Unternehmen...", len(alle_unternehmen))
        for i, u in enumerate(alle_unternehmen, 1):
            log.info("  [%d/%d] %s", i, len(alle_unternehmen), u.name)
            details_scrapen(u, session)

    return alle_unternehmen


def _scrape_mit_browser(
    max_seiten: int,
    branche: str,
    ort: str,
    score_min: str,
    details: bool,
) -> list[Unternehmen]:
    """Scrapt mit Playwright-Browser (rendert JavaScript)."""
    alle_unternehmen: list[Unternehmen] = []
    gesehene_urls: set[str] = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=HEADERS["User-Agent"],
            locale="de-DE",
        )
        page = context.new_page()

        for seite in range(1, max_seiten + 1):
            url = suchseite_url(seite, branche, ort, score_min)
            log.info("Seite %d/%d (Browser): %s", seite, max_seiten, url)

            soup = seite_abrufen_browser(url, page)
            if not soup:
                log.warning("Seite %d konnte nicht geladen werden.", seite)
                continue

            unternehmen = unternehmen_von_suchseite(soup)
            log.info("  %d Unternehmen auf Seite %d gefunden", len(unternehmen), seite)

            for u in unternehmen:
                if u.profil_url not in gesehene_urls:
                    gesehene_urls.add(u.profil_url)
                    alle_unternehmen.append(u)

            if not unternehmen:
                log.info("Keine weiteren Unternehmen gefunden, Abbruch.")
                break

            pause(2.0, 4.0)

        # Details mit Browser laden
        if details and alle_unternehmen:
            log.info("Lade Details für %d Unternehmen...", len(alle_unternehmen))
            session = requests.Session()
            session.headers.update(HEADERS)
            for i, u in enumerate(alle_unternehmen, 1):
                log.info("  [%d/%d] %s", i, len(alle_unternehmen), u.name)
                details_scrapen(u, session)

        browser.close()

    return alle_unternehmen


def als_csv_speichern(unternehmen: list[Unternehmen], dateiname: str) -> None:
    """Speichert die Daten als CSV-Datei."""
    if not unternehmen:
        log.warning("Keine Daten zum Speichern.")
        return

    feldnamen = [f.name for f in fields(Unternehmen)]

    with open(dateiname, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=feldnamen, delimiter=";")
        writer.writeheader()
        for u in unternehmen:
            writer.writerow(asdict(u))

    log.info("CSV gespeichert: %s (%d Einträge)", dateiname, len(unternehmen))


def als_xlsx_speichern(unternehmen: list[Unternehmen], dateiname: str) -> None:
    """Speichert die Daten als Excel-Datei."""
    if not unternehmen:
        log.warning("Keine Daten zum Speichern.")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        log.error(
            "openpyxl ist nicht installiert. "
            "Installiere es mit: pip install openpyxl"
        )
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "kununu Unternehmen"

    # Spaltenüberschriften (menschenlesbar)
    ueberschriften = {
        "name": "Unternehmen",
        "score": "Score",
        "anzahl_bewertungen": "Bewertungen",
        "standort": "Standort",
        "branche": "Branche",
        "profil_url": "Profil-URL",
        "weiterempfehlung": "Weiterempfehlung",
        "gehaltszufriedenheit": "Gehaltszufriedenheit",
        "karriere_weiterbildung": "Karriere/Weiterbildung",
        "unternehmenskultur": "Unternehmenskultur",
        "arbeitsumgebung": "Arbeitsumgebung",
        "vielfalt": "Vielfalt",
        "kollegenzusammenhalt": "Kollegenzusammenhalt",
        "kommunikation": "Kommunikation",
        "top_company": "Top Company",
    }

    feldnamen = [f.name for f in fields(Unternehmen)]

    # Header-Zeile formatieren
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, feld in enumerate(feldnamen, 1):
        cell = ws.cell(row=1, column=col_idx, value=ueberschriften.get(feld, feld))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Daten einfügen
    for row_idx, u in enumerate(unternehmen, 2):
        daten = asdict(u)
        for col_idx, feld in enumerate(feldnamen, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=daten[feld])
            cell.border = thin_border

            # URLs als Hyperlinks
            if feld == "profil_url" and daten[feld]:
                cell.hyperlink = daten[feld]
                cell.font = Font(color="0563C1", underline="single")

    # Spaltenbreiten anpassen
    spaltenbreiten = {
        "name": 35,
        "score": 10,
        "anzahl_bewertungen": 15,
        "standort": 25,
        "branche": 30,
        "profil_url": 45,
        "weiterempfehlung": 18,
        "gehaltszufriedenheit": 20,
        "karriere_weiterbildung": 22,
        "unternehmenskultur": 20,
        "arbeitsumgebung": 18,
        "vielfalt": 12,
        "kollegenzusammenhalt": 22,
        "kommunikation": 18,
        "top_company": 14,
    }

    for col_idx, feld in enumerate(feldnamen, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = spaltenbreiten.get(feld, 15)

    # Autofilter aktivieren
    ws.auto_filter.ref = ws.dimensions

    # Erste Zeile fixieren
    ws.freeze_panes = "A2"

    wb.save(dateiname)
    log.info("Excel gespeichert: %s (%d Einträge)", dateiname, len(unternehmen))


def main():
    parser = argparse.ArgumentParser(
        description="kununu.com Web Scraper - Unternehmensdaten extrahieren",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  python kununu_scraper.py
  python kununu_scraper.py --seiten 5 --format beide
  python kununu_scraper.py --seiten 3 --details --format xlsx
  python kununu_scraper.py --branche IT --ort Berlin
        """,
    )
    parser.add_argument(
        "--seiten", type=int, default=3,
        help="Anzahl der Suchergebnis-Seiten (Standard: 3)",
    )
    parser.add_argument(
        "--format", choices=["csv", "xlsx", "beide"], default="beide",
        help="Ausgabeformat (Standard: beide)",
    )
    parser.add_argument(
        "--ausgabe", type=str, default="kununu_unternehmen",
        help="Dateiname ohne Endung (Standard: kununu_unternehmen)",
    )
    parser.add_argument(
        "--details", action="store_true",
        help="Detailseiten der Unternehmen scrapen (langsamer)",
    )
    parser.add_argument(
        "--browser", action="store_true",
        help="Playwright-Browser nutzen (benötigt: pip install playwright)",
    )
    parser.add_argument(
        "--branche", type=str, default="",
        help="Nach Branche filtern",
    )
    parser.add_argument(
        "--ort", type=str, default="",
        help="Nach Ort filtern",
    )
    parser.add_argument(
        "--score", type=str, default="",
        help="Mindest-Score (z.B. '4' für 4-5 Sterne)",
    )

    args = parser.parse_args()

    log.info("=" * 60)
    log.info("kununu.com Web Scraper")
    log.info("=" * 60)
    log.info("Seiten: %d | Format: %s | Details: %s | Browser: %s",
             args.seiten, args.format,
             "Ja" if args.details else "Nein",
             "Ja" if args.browser else "Nein")
    if args.branche:
        log.info("Branche: %s", args.branche)
    if args.ort:
        log.info("Ort: %s", args.ort)
    log.info("-" * 60)

    # Scraping starten
    unternehmen = scrape(
        max_seiten=args.seiten,
        branche=args.branche,
        ort=args.ort,
        score_min=args.score,
        details=args.details,
        browser_modus=args.browser,
    )

    if not unternehmen:
        log.warning("Keine Unternehmen gefunden.")
        sys.exit(1)

    log.info("-" * 60)
    log.info("Insgesamt %d Unternehmen gefunden.", len(unternehmen))

    # Exportieren
    ausgabe = Path(args.ausgabe)

    if args.format in ("csv", "beide"):
        als_csv_speichern(unternehmen, str(ausgabe.with_suffix(".csv")))

    if args.format in ("xlsx", "beide"):
        als_xlsx_speichern(unternehmen, str(ausgabe.with_suffix(".xlsx")))

    log.info("Fertig!")


if __name__ == "__main__":
    main()
