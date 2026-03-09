#!/usr/bin/env python3
"""
kununu Bewertungen Scraper – Web-Frontend (Flask)

Starten mit:
    python app.py

Dann im Browser öffnen: http://localhost:5000
"""

import io
import csv
import os
import threading
import uuid
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file

app = Flask(__name__)

# Scraping-Jobs verfolgen: {job_id: {status, fortschritt, ergebnis, fehler}}
jobs: dict[str, dict] = {}


def _scrape_job(job_id: str, eingabe: str, fmt: str, typ: str):
    """Führt den Scraping-Job in einem Hintergrund-Thread aus."""
    from kununu_bewertungen_scraper import (
        PLAYWRIGHT_VERFUEGBAR, sync_playwright, firmen_url_finden,
        bewertungen_scrapen, MITARBEITER_FELDER, BEWERBER_FELDER,
    )

    job = jobs[job_id]

    if not PLAYWRIGHT_VERFUEGBAR:
        job["status"] = "fehler"
        job["fehler"] = "Playwright ist nicht installiert."
        return

    try:
        job["fortschritt"] = "Browser wird gestartet …"

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                locale="de-DE",
            )
            page = context.new_page()

            # URL oder Name?
            ist_url = eingabe.startswith("http://") or eingabe.startswith("https://")
            if ist_url:
                firmen_url = eingabe.rstrip("/")
            else:
                job["fortschritt"] = f"Suche '{eingabe}' auf kununu …"
                firmen_url = firmen_url_finden(eingabe, page)
                if not firmen_url:
                    job["status"] = "fehler"
                    job["fehler"] = f"Unternehmen '{eingabe}' nicht gefunden."
                    browser.close()
                    return

            slug = firmen_url.rstrip("/").split("/")[-1]
            job["firma"] = slug
            ergebnis = {}

            # Mitarbeiter
            if typ in ("beide", "mitarbeiter"):
                job["fortschritt"] = "Mitarbeiter-Bewertungen werden geladen …"
                mitarbeiter = bewertungen_scrapen(
                    firmen_url, "mitarbeiter", page,
                    max_seiten=50,
                )
                ergebnis["mitarbeiter"] = mitarbeiter
                job["fortschritt"] = f"{len(mitarbeiter)} Mitarbeiter-Bewertungen gefunden."

            # Bewerber
            if typ in ("beide", "bewerber"):
                job["fortschritt"] = "Bewerber-Bewertungen werden geladen …"
                bewerber = bewertungen_scrapen(
                    firmen_url, "bewerber", page,
                    max_seiten=50,
                )
                ergebnis["bewerber"] = bewerber
                job["fortschritt"] = f"{len(bewerber)} Bewerber-Bewertungen gefunden."

            browser.close()

        job["ergebnis"] = ergebnis
        job["format"] = fmt
        job["status"] = "fertig"

        # Zusammenfassung
        teile = []
        if "mitarbeiter" in ergebnis:
            teile.append(f"{len(ergebnis['mitarbeiter'])} Mitarbeiter")
        if "bewerber" in ergebnis:
            teile.append(f"{len(ergebnis['bewerber'])} Bewerber")
        job["fortschritt"] = f"Fertig! {', '.join(teile)}-Bewertungen gefunden."

    except Exception as e:
        job["status"] = "fehler"
        job["fehler"] = str(e)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/scrape", methods=["POST"])
def scrape_starten():
    data = request.get_json()
    eingabe = (data.get("eingabe") or "").strip()
    fmt = data.get("format", "csv")
    typ = data.get("typ", "beide")

    if not eingabe:
        return jsonify({"error": "Bitte einen Firmennamen oder URL eingeben."}), 400

    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "Ungültiges Format."}), 400

    if typ not in ("beide", "mitarbeiter", "bewerber"):
        return jsonify({"error": "Ungültiger Typ."}), 400

    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status": "laeuft",
        "fortschritt": "Wird gestartet …",
        "ergebnis": None,
        "fehler": None,
        "firma": "",
        "format": fmt,
    }

    thread = threading.Thread(target=_scrape_job, args=(job_id, eingabe, fmt, typ))
    thread.daemon = True
    thread.start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def job_status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job nicht gefunden."}), 404

    antwort = {
        "status": job["status"],
        "fortschritt": job["fortschritt"],
    }

    if job["status"] == "fehler":
        antwort["fehler"] = job["fehler"]

    if job["status"] == "fertig":
        antwort["firma"] = job["firma"]
        downloads = []
        if "mitarbeiter" in (job.get("ergebnis") or {}):
            downloads.append({
                "typ": "mitarbeiter",
                "label": f"Mitarbeiter ({len(job['ergebnis']['mitarbeiter'])})",
                "url": f"/api/download/{job_id}/mitarbeiter",
            })
        if "bewerber" in (job.get("ergebnis") or {}):
            downloads.append({
                "typ": "bewerber",
                "label": f"Bewerber ({len(job['ergebnis']['bewerber'])})",
                "url": f"/api/download/{job_id}/bewerber",
            })
        antwort["downloads"] = downloads

    return jsonify(antwort)


@app.route("/api/download/<job_id>/<typ>")
def download(job_id, typ):
    from kununu_bewertungen_scraper import MITARBEITER_FELDER, BEWERBER_FELDER

    job = jobs.get(job_id)
    if not job or job["status"] != "fertig":
        return jsonify({"error": "Job nicht verfügbar."}), 404

    if typ not in ("mitarbeiter", "bewerber"):
        return jsonify({"error": "Ungültiger Typ."}), 400

    daten = job["ergebnis"].get(typ, [])
    felder = MITARBEITER_FELDER if typ == "mitarbeiter" else BEWERBER_FELDER
    firma = job.get("firma", "export")
    dateiname_prefix = f"{firma}_{'employee_rows' if typ == 'mitarbeiter' else 'candidates_rows'}"
    fmt = job.get("format", "csv")

    if fmt == "xlsx":
        return _als_xlsx(daten, felder, f"{dateiname_prefix}.xlsx")
    else:
        return _als_csv(daten, felder, f"{dateiname_prefix}.csv")


def _als_csv(daten: list[dict], felder: list[str], dateiname: str):
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=felder, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(daten)
    mem = io.BytesIO(output.getvalue().encode("utf-8"))
    mem.seek(0)
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=dateiname)


def _als_xlsx(daten: list[dict], felder: list[str], dateiname: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Bewertungen"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, feld in enumerate(felder, 1):
        cell = ws.cell(row=1, column=col_idx, value=feld)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(daten, 2):
        for col_idx, feld in enumerate(felder, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(feld, ""))
            cell.border = thin_border

    for col_idx, feld in enumerate(felder, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(feld) + 4, 15)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=dateiname,
    )


if __name__ == "__main__":
    app.run(debug=False, port=5001)
